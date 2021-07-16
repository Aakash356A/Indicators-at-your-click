#stockcode 
import time
import datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
#import csv
#import googlefinance
#from pyexcel.cookbook import merge_all_to_a_book
#import glob

        

wrkbk = openpyxl.load_workbook("Stock_codes.xlsx")
sh = wrkbk.active
kf=0.15384 #k=smoothing factor ,k=2/(n+1), where n is number of days in EMA
k1=1-kf
k26=0.074074
k261=1-k26
k9=0.2
k91=1-k9



# iterate through excel and display data
for row in sh.iter_rows(min_row=1, min_col=1, max_row=1, max_col=1):
    for ticker in row:
        print(ticker.value, end=" ")
    print()
    
    ticker='DMART'
    period1 = int(time.mktime(datetime.datetime(1992, 1, 1, 23, 59).timetuple()))
    period2 = int(time.mktime(datetime.datetime(2021, 7, 28, 23, 59).timetuple()))
    interval = '1d' # 1d, 1m

    query_string =f'https://query1.finance.yahoo.com/v7/finance/download/{ticker}.NS?period1={period1}&period2={period2}&interval={interval}&events=history&includeAdjustedClose=true'

    d1=pd.read_csv(query_string)
    #print(d1)
    d2=pd.ExcelWriter('ticker.xlsx')
    d1.to_excel(d2,index=False)
    d2.save()

    
    #wrkbk1 = openpyxl.load_workbook("ticker.xlsx")
    #sh1 = wrkbk1.active
    
    coun=0
    excel_file=openpyxl.load_workbook('ticker.xlsx')
    excel_sheet=excel_file['Sheet1']
    sh1 = excel_file.active
    for row in excel_sheet:
        coun+=1
        for cell in row:
            if cell.value==None:
                excel_sheet.delete_rows(idx=coun)
    excel_file.save('ticker.xlsx')

    dategraph=[]
    '''
    for row1 in sh1.iter_rows(min_row=49,min_col=1,max_row=total1,max_col=1):
        for j in row1:
            #print(j.value)
            dategraph.append(j.value)
    #print(dategraph)
    '''
    sum1=0
    for row1 in sh1.iter_rows(min_row=2,min_col=5,max_row=13,max_col=5):
        for k in row1:
            #print(k.value)          
            sum1=sum1+ k.value    #avg of 1st 12
    EMA=sum1/12    
    #print(EMA)
    a26=[] 
    signal9=[]
    signalgraph=[]
    MACDgraph=[]
    a26.append(EMA)
    count=0
    count1=0
    total=sh1.max_row
    total1=int(total/100)
    total1=total-total1
    count2=-1
    for row2 in sh1.iter_rows(min_row=14,min_col=5,max_col=5):
        count=count+1
        
        for i in row2:
            l=i.value
            EMA=l*kf + EMA*k1
            #print(EMA)
            if count <26:
                a26.append(EMA) #avg of 26  
            elif count==26:
                    EMA26=sum(a26)/len(a26)
            elif count >26 :  #do for 26
                #print('okpoj')
                count1=count1+1
                EMA26=l*k26 + EMA26*k261
                #print(EMA26)
                MACD=EMA-EMA26
                #print('aakash')
                #MACDgraph.append(MACD)
                if count1<9:
                    signal9.append(MACD)
                    #print('Divakar')
                elif count1==9:
                    signal=sum(signal9)/len(signal9)
                    #print(signal)
                    #print('ad')
                elif count1>9:
                    if total-count<=50:
                        count2=count2+1
                        MACDgraph.append(MACD)
                        signalgraph.append(signal)
                    signal=MACD*k9 + signal*k91
                    #print(signal)
                    #print('afef')

    
    #print(total1)
    total2=total-count2
    for row1 in sh1.iter_rows(min_row=total2,min_col=1,max_col=1):
        for j in row1:
            #print(j.value)
            dategraph.append(j.value)
print(len(MACDgraph))
print(len(dategraph)) 
print(len(signalgraph))   

plt.plot(dategraph, MACDgraph, label = "MACD")
plt.plot(dategraph, signalgraph, label = "Signal")
plt.xlabel('Date')
plt.ylabel('Indicator')
plt.title('MACD Indicator')
plt.legend()
plt.show()

